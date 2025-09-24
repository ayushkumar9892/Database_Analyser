import getpass
import datetime
from typing import List, Dict, Optional, Tuple, Set
from difflib import SequenceMatcher
import re
import datetime

# Optional dependencies: make imports lazy-friendly so API can start without them
try:
	import openpyxl  # type: ignore
	from openpyxl import Workbook  # type: ignore
	from openpyxl.styles import Font, PatternFill, Alignment  # type: ignore
except Exception:  # pragma: no cover
	openpyxl = None  # type: ignore
	Workbook = None  # type: ignore
	Font = PatternFill = Alignment = None  # type: ignore

try:
	import psycopg2  # type: ignore
except Exception:  # pragma: no cover
	psycopg2 = None  # type: ignore

try:
	import pyodbc  # type: ignore
except Exception:  # pragma: no cover
	pyodbc = None  # type: ignore

try:
	import mysql.connector as mysql_connector  # type: ignore
except Exception:  # pragma: no cover
	mysql_connector = None  # type: ignore

import re


class DatabaseAnalyzer:
    """Enhanced Database Schema Analyzer with improved connection handling and features."""
    
    def __init__(self):
        self.conn = None
        self.cursor = None
        self.db_type = None
        
    def get_connection_params(self, db_type: str) -> Dict:
        """Get connection parameters with smart defaults and flexible input."""
        params = {}
        
        if db_type == "postgresql":
            params['host'] = input("Enter host (default: localhost): ").strip() or "localhost"
            params['port'] = input("Enter port (default: 5432): ").strip() or "5432"
            params['database'] = input("Enter database name: ").strip()
            params['username'] = input("Enter username: ").strip()
            params['password'] = getpass.getpass("Enter password (input hidden): ")
            
        elif db_type == "sqlserver":
            print("\nSQL Server Connection Options:")
            print("1. Server name only (Windows Authentication)")
            print("2. Server name with SQL Authentication")
            print("3. Full connection details")
            
            auth_choice = input("Choose option (1/2/3): ").strip()
            
            if auth_choice == "1":
                params['server'] = input("Enter server name: ").strip()
                params['database'] = input("Enter database name (default: master): ").strip() or "master"
                params['trusted_connection'] = True
                
            elif auth_choice == "2":
                params['server'] = input("Enter server name: ").strip()
                params['database'] = input("Enter database name (default: master): ").strip() or "master"
                params['username'] = input("Enter username: ").strip()
                params['password'] = getpass.getpass("Enter password (input hidden): ")
                params['trusted_connection'] = False
                
            else:  # Option 3
                params['host'] = input("Enter host: ").strip()
                params['port'] = input("Enter port (default: 1433): ").strip() or "1433"
                params['database'] = input("Enter database name: ").strip()
                params['username'] = input("Enter username: ").strip()
                params['password'] = getpass.getpass("Enter password (input hidden): ")
                params['trusted_connection'] = False
                
        elif db_type == "mysql":
            params['host'] = input("Enter host (default: localhost): ").strip() or "localhost"
            params['port'] = int(input("Enter port (default: 3306): ").strip() or "3306")
            params['database'] = input("Enter database name: ").strip()
            params['username'] = input("Enter username: ").strip()
            params['password'] = getpass.getpass("Enter password (input hidden): ")
            
        return params
    
    def connect_database(self, db_type: str, params: Dict) -> bool:
        """Establish database connection with enhanced error handling."""
        try:
            self.db_type = db_type
            
            if db_type == "postgresql":
				# Ensure driver is available
				if psycopg2 is None:
					raise ImportError("psycopg2-binary not installed")
				self.conn = psycopg2.connect(
                    host=params['host'],
                    port=params['port'],
                    dbname=params['database'],
                    user=params['username'],
                    password=params['password'],
                    connect_timeout=10
                )
                
            elif db_type == "sqlserver":
				# Ensure driver is available
				if pyodbc is None:
					raise ImportError("pyodbc not installed")
                if params.get('trusted_connection', False):
                    conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={params['server']};DATABASE={params['database']};Trusted_Connection=yes;"
                elif 'server' in params:
                    conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={params['server']};DATABASE={params['database']};UID={params['username']};PWD={params['password']}"
                else:
                    conn_str = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={params['host']},{params['port']};DATABASE={params['database']};UID={params['username']};PWD={params['password']}"
                
                self.conn = pyodbc.connect(conn_str, timeout=10)
                
            elif db_type == "mysql":
				# Ensure driver is available
				if mysql_connector is None:
					raise ImportError("mysql-connector-python not installed")
				self.conn = mysql_connector.connect(
                    host=params['host'],
                    port=params['port'],
                    database=params['database'],
                    user=params['username'],
                    password=params['password'],
                    connection_timeout=10
                )
                
            self.cursor = self.conn.cursor()
            print(f"✓ Successfully connected to {db_type.upper()} database!")
            return True
            
        except ImportError as e:
            print(f"✗ Required database driver not installed: {e}")
            print(f"Install it using: pip install {'psycopg2-binary' if db_type == 'postgresql' else 'pyodbc' if db_type == 'sqlserver' else 'mysql-connector-python'}")
            return False
        except Exception as e:
            print(f"✗ Connection failed: {e}")
            return False
          

    def get_tables(self) -> List[Tuple[str, str]]:
        """Fetches and returns a list of (schema, table_name) pairs."""
        try:
            if self.db_type == "mysql":
                self.cursor.execute("""
                    SELECT TABLE_SCHEMA, TABLE_NAME 
                    FROM INFORMATION_SCHEMA.TABLES 
                    WHERE TABLE_TYPE = 'BASE TABLE'
                """)
                tables = [(row[0], row[1]) for row in self.cursor.fetchall()]

            elif self.db_type == "sqlserver":
                self.cursor.execute("""
                    SELECT TABLE_SCHEMA, TABLE_NAME 
                    FROM INFORMATION_SCHEMA.TABLES 
                    WHERE TABLE_TYPE = 'BASE TABLE'
                """)
                tables = [(row[0], row[1]) for row in self.cursor.fetchall()]

            elif self.db_type == "postgresql":
                self.cursor.execute("""
                    SELECT table_schema, table_name 
                    FROM information_schema.tables 
                    WHERE table_type='BASE TABLE'
                """)
                tables = [(row[0], row[1]) for row in self.cursor.fetchall()]

            return sorted(tables)

        except Exception as e:
            print(f"✗ Error fetching tables: {e}")
            return []


    def get_views(self) -> List[Tuple[str, str]]:
        """Fetches and returns a list of (schema, view_name) pairs."""
        try:
            if self.db_type.lower() == "mysql":
                self.cursor.execute("""
                    SELECT TABLE_SCHEMA, TABLE_NAME
                    FROM INFORMATION_SCHEMA.TABLES
                    WHERE TABLE_TYPE = 'VIEW'
                    AND TABLE_SCHEMA NOT IN ('mysql', 'information_schema', 'performance_schema', 'sys')
                """)
                views = [(row[0], row[1]) for row in self.cursor.fetchall()]

            elif self.db_type.lower() == "sqlserver":
                self.cursor.execute("""
                    SELECT TABLE_SCHEMA, TABLE_NAME
                    FROM INFORMATION_SCHEMA.VIEWS
                """)
                views = [(row[0], row[1]) for row in self.cursor.fetchall()]

            elif self.db_type.lower() == "postgresql":
                # exclude system schemas
                self.cursor.execute("""
                    SELECT table_schema, table_name
                    FROM information_schema.views
                    WHERE table_schema NOT IN ('pg_catalog', 'information_schema')
                """)
                views = [(row[0], row[1]) for row in self.cursor.fetchall()]

            else:
                raise ValueError(f"Unsupported db_type: {self.db_type}")

            return sorted(views)

        except Exception as e:
            print(f"✗ Error fetching views: {e}")
            return []



    def export_all_tables_analysis(self):
        """Export detailed analysis of all tables to Excel."""
        print(f"\n{'='*20} EXPORT ALL TABLES ANALYSIS {'='*20}")
        
        tables = self.get_tables()
        if not tables:
            print("No tables found to analyze.")
            return
        
        print(f"Found {len(tables)} tables to analyze and export...")
        
        # Confirm if many tables
        if len(tables) > 20:
            proceed = input(f"This will analyze {len(tables)} tables and may take time. Continue? (y/N): ").strip().lower()
            if proceed != 'y':
                return
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "All Tables Analysis"
            
            # Add header
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws['A1'] = f"All Tables Analysis Report - {self.db_type.upper()}"
            ws['A1'].font = Font(size=14, bold=True)
            ws['A2'] = f"Generated: {timestamp}"
            ws['A3'] = f"Total Tables: {len(tables)}"
            
            current_row = 5
            
            # Process each table
            for table_idx, (schema, table_name) in enumerate(tables, 1):
                print(f"Processing {table_idx}/{len(tables)}: {schema}.{table_name}")
                
                try:
                    # Table header
                    ws[f'A{current_row}'] = f"TABLE: {schema}.{table_name}"
                    ws[f'A{current_row}'].font = Font(size=12, bold=True)
                    ws[f'A{current_row}'].fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                    ws[f'A{current_row}'].font = Font(color="FFFFFF", bold=True)
                    current_row += 1
                    
                    # Basic table stats
                    self.cursor.execute(f"SELECT COUNT(*) FROM {schema}.{table_name}")
                    row_count = self.cursor.fetchone()[0]
                    ws[f'A{current_row}'] = f"Total Rows: {row_count:,}"
                    current_row += 1
                    
                    # Column information
                    columns_info = self._get_column_info(schema, table_name)
                    ws[f'A{current_row}'] = f"Total Columns: {len(columns_info)}"
                    current_row += 1
                    
                    # Table size
                    table_size = self._get_table_size(schema, table_name)
                    if table_size:
                        ws[f'A{current_row}'] = f"Estimated Size: {table_size}"
                        current_row += 1
                    
                    # Column details header
                    current_row += 1
                    ws[f'A{current_row}'] = "Column Name"
                    ws[f'B{current_row}'] = "Data Type"
                    ws[f'C{current_row}'] = "Nullable"
                    ws[f'D{current_row}'] = "Default"
                    
                    # Style column headers
                    for col in ['A', 'B', 'C', 'D']:
                        cell = ws[f'{col}{current_row}']
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                    
                    current_row += 1
                    
                    # Column details
                    for col_info in columns_info:
                        ws[f'A{current_row}'] = col_info['name']
                        ws[f'B{current_row}'] = col_info['type']
                        ws[f'C{current_row}'] = col_info['nullable']
                        ws[f'D{current_row}'] = str(col_info.get('default', 'None'))[:30]
                        current_row += 1
                    
                    # Data quality analysis (for reasonable-sized tables)
                    if 0 < row_count <= 100000:
                        current_row += 1
                        ws[f'A{current_row}'] = "Data Quality Analysis"
                        ws[f'A{current_row}'].font = Font(bold=True)
                        ws[f'A{current_row}'].fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
                        current_row += 1
                        
                        quality_issues = []
                        
                        for col_info in columns_info:
                            col_name = col_info['name']
                            is_nullable = col_info['nullable'] == 'YES'
                            
                            try:
                                # Check for NULL values
                                self.cursor.execute(f"SELECT COUNT(*) FROM {schema}.{table_name} WHERE {col_name} IS NULL")
                                null_count = self.cursor.fetchone()[0]
                                
                                # Check for empty strings (for string columns)
                                empty_count = 0
                                if 'char' in col_info['type'].lower() or 'text' in col_info['type'].lower():
                                    self.cursor.execute(f"SELECT COUNT(*) FROM {schema}.{table_name} WHERE {col_name} = ''")
                                    empty_count = self.cursor.fetchone()[0]
                                
                                # Get distinct count
                                self.cursor.execute(f"SELECT COUNT(DISTINCT {col_name}) FROM {schema}.{table_name}")
                                distinct_count = self.cursor.fetchone()[0]
                                
                                # Calculate percentages
                                null_pct = (null_count / row_count) * 100 if row_count > 0 else 0
                                empty_pct = (empty_count / row_count) * 100 if row_count > 0 else 0
                                cardinality = (distinct_count / row_count) * 100 if row_count > 0 else 0
                                
                                # Report findings
                                status = "OK"
                                notes = []
                                
                                if not is_nullable and null_count > 0:
                                    status = "WARNING"
                                    notes.append(f"{null_count} NULLs in non-nullable column")
                                    quality_issues.append(f"Column '{col_name}': NULL constraint violation")
                                
                                if null_pct > 50:
                                    status = "WARNING"
                                    notes.append(f"High NULL rate: {null_pct:.1f}%")
                                
                                if empty_pct > 20:
                                    status = "WARNING"
                                    notes.append(f"High empty string rate: {empty_pct:.1f}%")
                                
                                if cardinality < 1 and row_count > 1:
                                    status = "WARNING"
                                    notes.append("All values identical")
                                
                                note_text = "; ".join(notes) if notes else f"Distinct: {distinct_count} ({cardinality:.1f}%)"
                                
                                ws[f'A{current_row}'] = f"{status}: {col_name}"
                                ws[f'B{current_row}'] = note_text
                                
                                # Color code warnings
                                if status == "WARNING":
                                    ws[f'A{current_row}'].fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                                    ws[f'B{current_row}'].fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                                
                                current_row += 1
                                
                            except Exception as e:
                                ws[f'A{current_row}'] = f"ERROR: {col_name}"
                                ws[f'B{current_row}'] = f"Error analyzing: {e}"
                                ws[f'A{current_row}'].fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                                ws[f'B{current_row}'].fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                                current_row += 1
                        
                        if quality_issues:
                            current_row += 1
                            ws[f'A{current_row}'] = "Quality Issues Summary:"
                            ws[f'A{current_row}'].font = Font(bold=True)
                            current_row += 1
                            for issue in quality_issues:
                                ws[f'A{current_row}'] = f"- {issue}"
                                current_row += 1
                    else:
                        current_row += 1
                        skip_reason = "Table too large" if row_count > 100000 else "Table empty"
                        ws[f'A{current_row}'] = f"Data Quality Analysis skipped: {skip_reason}"
                        ws[f'A{current_row}'].fill = PatternFill(start_color="FFF2E6", end_color="FFF2E6", fill_type="solid")
                        current_row += 1
                    
                    # Add separator between tables
                    current_row += 2
                    
                except Exception as e:
                    ws[f'A{current_row}'] = f"ERROR analyzing table '{schema}.{table_name}': {e}"
                    ws[f'A{current_row}'].fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                    current_row += 2
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 80)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save file
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"all_tables_analysis_{self.db_type}_{timestamp}.xlsx"
            wb.save(filename)
            
            print(f"\nAll tables analysis exported to: {filename}")
            print(f"Analysis completed for {len(tables)} tables")
            


    def find_tables_by_column(self, column_name: str):
        """Find all tables that contain a specific column name."""
        print(f"\n{'='*20} TABLES CONTAINING COLUMN: '{column_name}' {'='*20}")
        
        try:
            matching_tables = []
            
            if self.db_type == "mysql":
                self.cursor.execute("""
                    SELECT TABLE_SCHEMA, TABLE_NAME, COLUMN_TYPE, IS_NULLABLE, COLUMN_DEFAULT
                    FROM INFORMATION_SCHEMA.COLUMNS
                    WHERE COLUMN_NAME = %s
                    ORDER BY TABLE_SCHEMA, TABLE_NAME
                """, (column_name,))
                
            elif self.db_type == "postgresql":
                self.cursor.execute("""
                    SELECT table_schema, table_name, data_type, is_nullable, column_default
                    FROM information_schema.columns
                    WHERE column_name = %s
                    ORDER BY table_schema, table_name
                """, (column_name,))
                
            elif self.db_type == "sqlserver":
                self.cursor.execute("""
                    SELECT TABLE_SCHEMA, TABLE_NAME, DATA_TYPE, IS_NULLABLE, COLUMN_DEFAULT
                    FROM INFORMATION_SCHEMA.COLUMNS
                    WHERE COLUMN_NAME = ?
                    ORDER BY TABLE_SCHEMA, TABLE_NAME
                """, (column_name,))
            
            results = self.cursor.fetchall()
            
            if not results:
                print(f"No tables found containing column '{column_name}'")
                return []
            
            print(f"Found {len(results)} tables containing column '{column_name}':")
            print("-" * 80)
            print(f"{'No.':<4} {'Schema.Table':<35} {'Data Type':<20} {'Nullable':<10} {'Default':<15}")
            print("-" * 80)
            
            for i, row in enumerate(results, 1):
                schema = row[0]
                table_name = row[1]
                data_type = row[2]
                nullable = row[3]
                default_val = str(row[4])[:14] if row[4] else 'None'
                
                qualified_name = f"{schema}.{table_name}"
                matching_tables.append({
                    'schema': schema,
                    'table': table_name,
                    'qualified_name': qualified_name,
                    'data_type': data_type,
                    'nullable': nullable,
                    'default': row[4]
                })
                
                print(f"{i:<4} {qualified_name:<35} {data_type:<20} {nullable:<10} {default_val:<15}")
            
            # Offer to export results
            if len(results) > 5:
                export_choice = input(f"\nExport results to Excel? (y/N): ").strip().lower()
                if export_choice == 'y':
                    self._export_column_search_to_excel(column_name, matching_tables)
            
            return matching_tables
            
        except Exception as e:
            print(f"Error searching for column '{column_name}': {e}")
            return []


    def _export_column_search_to_excel(self, column_name: str, matching_tables: List[Dict]):
        """Export column search results to Excel file."""
        try:
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Column Search Results"
            
            # Header
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws['A1'] = f"Column Search Results: '{column_name}'"
            ws['A1'].font = Font(size=14, bold=True)
            ws['A2'] = f"Generated: {timestamp}"
            ws['A3'] = f"Database Type: {self.db_type.upper()}"
            ws['A4'] = f"Tables Found: {len(matching_tables)}"
            
            # Column headers
            row_num = 6
            headers = ['Schema', 'Table Name', 'Qualified Name', 'Data Type', 'Nullable', 'Default Value']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            
            # Data rows
            row_num += 1
            for table_info in matching_tables:
                ws.cell(row=row_num, column=1, value=table_info['schema'])
                ws.cell(row=row_num, column=2, value=table_info['table'])
                ws.cell(row=row_num, column=3, value=table_info['qualified_name'])
                ws.cell(row=row_num, column=4, value=table_info['data_type'])
                ws.cell(row=row_num, column=5, value=table_info['nullable'])
                ws.cell(row=row_num, column=6, value=str(table_info['default']) if table_info['default'] else 'None')
                row_num += 1
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save file
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"column_search_{column_name}_{timestamp}.xlsx"
            wb.save(filename)
            
            print(f"Column search results exported to: {filename}")
            
        except Exception as e:
            print(f"Error exporting to Excel: {e}")



    def detect_similar_tables(self):
        """Detect similar tables based on naming patterns and fuzzy matching."""
        print(f"\n{'='*20} SIMILAR TABLES DETECTION {'='*20}")
        
        try:
            tables = self.get_tables()
            # tables = [table_name for _, table_name in self.get_tables()]

            if len(tables) < 2:
                print("Need at least 2 tables to detect similarities.")
                return
            
            print(f"Analyzing {len(tables)} tables for similarities...")
            
            similar_groups = []
            
            fuzzy_groups = self._find_fuzzy_similar_tables(tables)
            similar_groups.extend(fuzzy_groups)
            
            if similar_groups:
                # Enhance with additional table information
                enhanced_groups = self._enhance_similar_tables_with_info(similar_groups)
                self._display_similar_tables_results(enhanced_groups)
                
                # Offer export
                export_choice = input(f"\nExport {len(enhanced_groups)} similar table groups to Excel? (y/N): ").strip().lower()
                if export_choice == 'y':
                    self._export_similar_tables_to_excel(enhanced_groups)
            else:
                print("No similar tables found based on the selected criteria.")
                
        except Exception as e:
            print(f"Error detecting similar tables: {e}")


    def _find_fuzzy_similar_tables(self, tables: List[Tuple[str, str]], similarity_threshold: float = 0.7) -> List[Dict]:
        """Find similar tables using fuzzy string matching."""
        print(f"Performing fuzzy matching (threshold: {similarity_threshold*100:.0f}%)...")
        
        # Get similarity threshold from user
        threshold_input = input(f"Enter similarity threshold (0.5-0.9, default {similarity_threshold}): ").strip()
        if threshold_input:
            try:
                similarity_threshold = float(threshold_input)
                similarity_threshold = max(0.5, min(0.9, similarity_threshold))
            except ValueError:
                pass
        
        processed = set()
        similar_groups = []
        
        for i, (schema1, table1) in enumerate(tables):
            if schema1 + '.' + table1 in processed:
                continue
            
            similar_group = [schema1 + '.' + table1]
            processed.add(schema1 + '.' + table1)
            
            for j, (schema2, table2) in enumerate(tables[i+1:], i+1):
                if schema2 + '.' + table2 in processed:
                    continue
                
                similarity = self._calculate_table_similarity(schema1 + '.' + table1, schema2 + '.' + table2)
                if similarity >= similarity_threshold:
                    similar_group.append(schema2 + '.' + table2)
                    processed.add(schema2 + '.' + table2)
            
            if len(similar_group) > 1:
                similar_groups.append({
                    'type': 'fuzzy',
                    'tables': similar_group,
                    'similarity': similarity_threshold * 100,
                    'group_size': len(similar_group)
                })
        
        return similar_groups


    def _calculate_table_similarity(self, table1: str, table2: str) -> float:
        """Calculate similarity between two table names."""
        if not table1 or not table2:
            return 0.0
        
        # Normalize names
        name1 = table1.lower().strip()
        name2 = table2.lower().strip()
        
        if name1 == name2:
            return 1.0
        
        # Use SequenceMatcher for basic similarity
        basic_similarity = SequenceMatcher(None, name1, name2).ratio()
        
        # Check for common prefixes/suffixes
        prefix_bonus = 0
        suffix_bonus = 0
        
        # Common prefix
        common_prefix = 0
        min_len = min(len(name1), len(name2))
        for i in range(min_len):
            if name1[i] == name2[i]:
                common_prefix += 1
            else:
                break
        
        if common_prefix > 3:  # At least 4 characters match from start
            prefix_bonus = 0.1
        
        # Check if one name is contained in another
        if name1 in name2 or name2 in name1:
            suffix_bonus = 0.15
        
        # Check for common words
        words1 = set(re.findall(r'[a-zA-Z]+', name1))
        words2 = set(re.findall(r'[a-zA-Z]+', name2))
        
        if words1 and words2:
            word_similarity = len(words1.intersection(words2)) / len(words1.union(words2))
            if word_similarity > 0.5:
                suffix_bonus += 0.1
        
        final_similarity = basic_similarity + prefix_bonus + suffix_bonus
        return min(1.0, final_similarity)

    def _enhance_similar_tables_with_info(self, similar_groups: List[Dict]) -> List[Dict]:
        """Enhance similar table groups with additional metadata."""
        enhanced_groups = []
        
        for group in similar_groups:
            enhanced_group = group.copy()
            table_info = {}
            
            for table in group['tables']:
                try:
                    schema = table.split('.')[0]
                    # Get row count
                    self.cursor.execute(f"SELECT COUNT(*) FROM {table}")
                    row_count = self.cursor.fetchone()[0]
                    
                    # Get table size if possible
                    table_size = self._get_table_size(schema, table)
                    
                    # Get creation date if possible
                    creation_date = self._get_table_creation_date(schema, table)
                    
                    # Get column count
                    columns_info = self._get_column_info(schema, table)
                    column_count = len(columns_info)
                    
                    table_info[table] = {
                        'row_count': row_count,
                        'size': table_size,
                        'creation_date': creation_date,
                        'column_count': column_count
                    }
                    
                except Exception as e:
                    table_info[table] = {
                        'row_count': 'Error',
                        'size': 'Error',
                        'creation_date': 'Error',
                        'column_count': 'Error',
                        'error': str(e)
                    }
            
            enhanced_group['table_info'] = table_info
            enhanced_groups.append(enhanced_group)
        
        return enhanced_groups

    def _get_table_creation_date(self, schema: str, table_name: str) -> Optional[str]:
        """Get table creation date if available."""
        try:
            if self.db_type == "sqlserver":
                self.cursor.execute(f"""
                    SELECT create_date 
                    FROM sys.tables t
                    JOIN sys.schemas s ON t.schema_id = s.schema_id
                    WHERE s.name = '{schema}' and t.name = '{table_name}'
                """)
                result = self.cursor.fetchone()
                return str(result[0]) if result else None
                
            elif self.db_type == "mysql":
                self.cursor.execute(f"""
                    SELECT CREATE_TIME 
                    FROM information_schema.TABLES 
                    WHERE TABLE_SCHEMA = '{schema}' AND TABLE_NAME = '{table_name}'
                """)
                result = self.cursor.fetchone()
                return str(result[0]) if result and result[0] else None
                
            elif self.db_type == "postgresql":
                # PostgreSQL doesn't store creation date, return None
                return None
                
        except Exception:
            return None

    

    def _display_similar_tables_results(self, enhanced_groups: List[Dict]):
        """Display the results of similar table detection."""
        print(f"\nSIMILAR TABLES DETECTION RESULTS:")
        print("=" * 70)
        
        total_tables = sum(group['group_size'] for group in enhanced_groups)
        print(f"Found {len(enhanced_groups)} similar table groups involving {total_tables} tables")
        
        for i, group in enumerate(enhanced_groups, 1):
            print(f"\nGroup {i}: {group['type'].upper()} similarity")
            print("-" * 50)
            
            print(f"No. of tables in group {i} -> ({group['group_size']}):")
            
            # Sort tables by risk level for better display
            tables_with_info = []
            for table in group['tables']:
                info = group['table_info'].get(table, {})
                tables_with_info.append((table, info))
            
            
            for table, info in tables_with_info:
                row_count = info.get('row_count', 'N/A')
                
                print(f"  🟡  {table}")
                print(f"    - Rows: {row_count:,}" if isinstance(row_count, int) else f"    - Rows: {row_count}")
                
                if 'error' in info:
                    print(f"    - Error: {info['error']}")

    def _export_similar_tables_to_excel(self, enhanced_groups: List[Dict]):
        """Export similar tables analysis to Excel."""
        try:
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Similar Tables Summary"
            
            # Header
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws['A1'] = f"Similar Tables Analysis Report"
            ws['A1'].font = Font(size=14, bold=True)
            ws['A2'] = f"Generated: {timestamp}"
            ws['A3'] = f"Database Type: {self.db_type.upper()}"
            
            # Summary
            row_num = 5
            headers = ['Group', 'Type', 'Table Count', 'Table Name', 'Row Count']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            
            row_num += 1
            
            # Data rows
            for group_num, group in enumerate(enhanced_groups, 1):
                for table_num, table in enumerate(group['tables']):
                    info = group['table_info'].get(table, {})
                    
                    ws.cell(row=row_num, column=1, value=group_num if table_num == 0 else "")
                    ws.cell(row=row_num, column=2, value=group['type'].upper() if table_num == 0 else "")
                    ws.cell(row=row_num, column=3, value=group['group_size'] if table_num == 0 else "")
                    ws.cell(row=row_num, column=4, value=table)
                    ws.cell(row=row_num, column=5, value=info.get('row_count', 'N/A'))
                    
                    row_num += 1
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save file
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"similar_tables_analysis_{timestamp}.xlsx"
            wb.save(filename)
            
            print(f"Similar tables analysis exported to: {filename}")
            
        except Exception as e:
            print(f"Error exporting to Excel: {e}")

    def close_connection(self):
        """Close database connection."""
        try:
            if self.cursor:
                self.cursor.close()
            if self.conn:
                self.conn.close()
            print("Database connection closed.")
        except Exception as e:
            print(f"Error closing connection: {e}")

    
    def get_database_overview(self):
        """Get comprehensive database overview."""
        print("\n" + "="*60)
        print(f"DATABASE OVERVIEW - {self.db_type.upper()}")
        print("="*60)
        
        try:
            # Tables count
            tables = self.get_tables()
            print(f"📋 Total Tables: {len(tables)}")
            
            # Views count
            views_count = self._get_views_count()
            print(f"👁  Total Views: {views_count}")
            
            # Schemas count
            schemas_count = self._get_schemas_count()
            print(f"🏗  Total Schemas: {schemas_count}")
            
            # Database size (if available)
            db_size = self._get_database_size()
            if db_size:
                print(f"💾 Database Size: {db_size}")
                
        except Exception as e:
            print(f"✗ Error getting database overview: {e}")
    
    def _get_views_count(self) -> int:
        """Get count of views."""
        try:
            if self.db_type == "mysql":
                self.cursor.execute("SELECT COUNT(*) FROM information_schema.views WHERE table_schema = DATABASE()")
            elif self.db_type == "sqlserver":
                self.cursor.execute("SELECT COUNT(*) FROM INFORMATION_SCHEMA.VIEWS WHERE TABLE_SCHEMA = 'dbo'")
            elif self.db_type == "postgresql":
                self.cursor.execute("SELECT COUNT(*) FROM information_schema.views WHERE table_schema='public'")
            return self.cursor.fetchone()[0]
        except:
            return 0
    
    def _get_schemas_count(self) -> int:
        """Get count of schemas."""
        try:
            if self.db_type == "mysql":
                self.cursor.execute("SELECT COUNT(*) FROM information_schema.schemata")
            elif self.db_type == "postgresql":
                self.cursor.execute("SELECT COUNT(*) FROM information_schema.schemata")
            elif self.db_type == "sqlserver":
                self.cursor.execute("SELECT COUNT(*) FROM sys.schemas")
            return self.cursor.fetchone()[0]
        except:
            return 0
    
    def _get_database_size(self) -> Optional[str]:
        """Get database size if available."""
        try:
            if self.db_type == "postgresql":
                self.cursor.execute("SELECT pg_size_pretty(pg_database_size(current_database()))")
                return self.cursor.fetchone()[0]
            elif self.db_type == "mysql":
                self.cursor.execute("""
                    SELECT ROUND(SUM(data_length + index_length) / 1024 / 1024, 2) AS 'DB Size in MB'
                    FROM information_schema.tables 
                    WHERE table_schema = DATABASE()
                """)
                result = self.cursor.fetchone()[0]
                return f"{result} MB" if result else None
            elif self.db_type == "sqlserver":
                self.cursor.execute("""
                    SELECT 
                        CAST(SUM(CAST(FILEPROPERTY(name, 'SpaceUsed') AS bigint) * 8192.) / 1024 / 1024 AS DECIMAL(15,2)) as 'DatabaseSizeInMB'
                    FROM sys.database_files
                    WHERE type_desc='ROWS'
                """)
                result = self.cursor.fetchone()[0]
                return f"{result} MB" if result else None
        except:
            return None
    
    def list_indexes_for_table(self, schema: str, table_name: str):
        """Lists indexes for a specific table with enhanced formatting."""
        print(f"\n{'='*20} INDEXES FOR TABLE: {schema}.{table_name} {'='*20}")
        
        try:
            if self.db_type == "mysql":
                self.cursor.execute(f"SHOW INDEXES FROM `{schema}.{table_name}`")
                indexes = self.cursor.fetchall()
                if not indexes:
                    print("ℹ️  No indexes found for this table.")
                    return

                print(f"{'Index Name':<25} {'Column Name':<25} {'Unique':<8} {'Type':<10}")
                print("-" * 75)
                for index_info in indexes:
                    index_name = index_info[2]
                    column_name = index_info[4]
                    non_unique = index_info[1]
                    is_unique = "Yes" if non_unique == 0 else "No"
                    index_type = index_info[10] if len(index_info) > 10 else "BTREE"
                    print(f"{index_name:<25} {column_name:<25} {is_unique:<8} {index_type:<10}")

            elif self.db_type == "postgresql":
                self.cursor.execute(f"""
                    SELECT
                        i.indexname,
                        i.indexdef,
                        ix.indisunique,
                        ix.indisprimary
                    FROM
                        pg_indexes i
                    JOIN pg_class t ON t.relname = i.tablename
                    JOIN pg_index ix ON ix.indexrelid = (
                        SELECT oid FROM pg_class WHERE relname = i.indexname
                    )
                    WHERE
                        i.schemaname = '{schema}' AND i.tablename = '{table_name}';
                """)
                indexes = self.cursor.fetchall()
                if not indexes:
                    print("ℹ️  No indexes found for this table.")
                    return

                print(f"{'Index Name':<25} {'Type':<10} {'Unique':<8} {'Primary':<8}")
                print("-" * 60)
                for index_info in indexes:
                    index_name = index_info[0]
                    is_unique = "Yes" if index_info[2] else "No"
                    is_primary = "Yes" if index_info[3] else "No"
                    index_type = "BTREE"  # Default for PostgreSQL
                    print(f"{index_name:<25} {index_type:<10} {is_unique:<8} {is_primary:<8}")
                    
            elif self.db_type == "sqlserver":
                self.cursor.execute(f"""
                    SELECT
                        ind.name AS IndexName,
                        STRING_AGG(COL_NAME(ic.object_id, ic.column_id), ', ') AS ColumnNames,
                        ind.is_unique,
                        ind.is_primary_key,
                        ind.type_desc
                    FROM
                        sys.indexes ind
                    JOIN
                        sys.index_columns ic ON ind.object_id = ic.object_id AND ind.index_id = ic.index_id
                    JOIN
                        sys.tables t ON ind.object_id = t.object_id                    
                    JOIN
                        sys.schemas s ON t.schema_id = s.schema_id
                    WHERE
                        s.name = '{schema}' AND t.name = '{table_name}' AND t.is_ms_shipped = 0
                    GROUP BY
                        ind.name, ind.is_unique, ind.is_primary_key, ind.type_desc
                    ORDER BY
                        ind.name;
                """)
                indexes = self.cursor.fetchall()
                if not indexes:
                    print("ℹ️  No indexes found for this table.")
                    return

                print(f"{'Index Name':<25} {'Columns':<30} {'Unique':<8} {'Primary':<8} {'Type':<12}")
                print("-" * 90)
                for row in indexes:
                    index_name = row[0]
                    columns = row[1]
                    is_unique = "Yes" if row[2] else "No"
                    is_primary = "Yes" if row[3] else "No"
                    index_type = row[4]
                    print(f"{index_name:<25} {columns:<30} {is_unique:<8} {is_primary:<8} {index_type:<12}")

        except Exception as e:
            print(f"✗ Error listing indexes for {schema}.{table_name}: {e}")
    
    def get_table_details_and_quality(self, schema: str, table_name: str):
        """Retrieves detailed information about a table with enhanced analysis."""
        print(f"\n{'='*20} TABLE ANALYSIS: {schema}.{table_name} {'='*20}")
        
        try:
            # 1. Basic table stats
            self.cursor.execute(f"SELECT COUNT(*) FROM {schema}.{table_name}")
            row_count = self.cursor.fetchone()[0]
            print(f"📊 Total Rows: {row_count:,}")

            # 2. Column information
            columns_info = self._get_column_info(schema, table_name)
            print(f"📋 Total Columns: {len(columns_info)}")
            
            # 3. Table size estimation
            table_size = self._get_table_size(schema, table_name)
            if table_size:
                print(f"💾 Estimated Size: {table_size}")
            
            print(f"\n{'Column Details':<60}")
            print("-" * 80)
            print(f"{'Column Name':<25} {'Data Type':<20} {'Nullable':<10} {'Default':<15}")
            print("-" * 80)
            
            for col_info in columns_info:
                default_val = str(col_info.get('default', 'None'))[:14]
                print(f"{col_info['name']:<25} {col_info['type']:<20} {col_info['nullable']:<10} {default_val:<15}")

            # 4. Data quality analysis
            if row_count > 0 and row_count <= 1000000:  # Only for reasonable-sized tables
                print(f"\n{'Data Quality Analysis':<60}")
                print("-" * 80)
                self._analyze_data_quality(table_name, columns_info, row_count)
            else:
                print(f"\n⚠️  Skipping data quality analysis (table too large: {row_count:,} rows)")

        except Exception as e:
            print(f"✗ Error analyzing table '{table_name}': {e}")
    

    def _get_column_info(self, schema: str, table_name: str) -> List[Dict]:
        """Get detailed column information for a given schema and table."""
        columns_info = []

        if self.db_type == "mysql":
            self.cursor.execute(f"""
                SELECT COLUMN_NAME, COLUMN_TYPE, IS_NULLABLE, COLUMN_DEFAULT
                FROM INFORMATION_SCHEMA.COLUMNS
                WHERE TABLE_SCHEMA = %s AND TABLE_NAME = %s
                ORDER BY ORDINAL_POSITION
            """, (schema, table_name))
            columns_data = self.cursor.fetchall()
            for col in columns_data:
                columns_info.append({
                    'name': col[0],
                    'type': col[1],
                    'nullable': col[2],
                    'default': col[3]
                })

        elif self.db_type == "postgresql":
            self.cursor.execute(f"""
                SELECT column_name, data_type, is_nullable, column_default
                FROM information_schema.columns
                WHERE table_schema = %s AND table_name = %s
                ORDER BY ordinal_position
            """, (schema, table_name))
            columns_data = self.cursor.fetchall()
            for col in columns_data:
                columns_info.append({
                    'name': col[0],
                    'type': col[1],
                    'nullable': col[2],
                    'default': col[3]
                })

        elif self.db_type == "sqlserver":
            self.cursor.execute(f"""
                SELECT COLUMN_NAME, DATA_TYPE, IS_NULLABLE, COLUMN_DEFAULT
                FROM INFORMATION_SCHEMA.COLUMNS
                WHERE TABLE_SCHEMA = ? AND TABLE_NAME = ?
                ORDER BY ORDINAL_POSITION
            """, (schema, table_name))
            columns_data = self.cursor.fetchall()
            for col in columns_data:
                columns_info.append({
                    'name': col[0],
                    'type': col[1],
                    'nullable': col[2],
                    'default': col[3]
                })

        return columns_info

    
    def _get_table_size(self, schema: str, table_name: str) -> Optional[str]:
        """Get table size estimation."""
        try:
            if self.db_type == "postgresql":
                self.cursor.execute(f"SELECT pg_size_pretty(pg_total_relation_size('{schema}.{table_name}'))")
                return self.cursor.fetchone()[0]
            elif self.db_type == "mysql":
                self.cursor.execute(f"""
                    SELECT ROUND(((data_length + index_length) / 1024 / 1024), 2) AS 'Size in MB'
                    FROM information_schema.TABLES 
                    WHERE table_schema = DATABASE() AND table_name = '{schema}.{table_name}'
                """)
                result = self.cursor.fetchone()[0]
                return f"{result} MB" if result else None
        except:
            return None
    
    def _analyze_data_quality(self, table_name: str, columns_info: List[Dict], row_count: int):
        """Perform basic data quality analysis."""
        quality_issues = []
        
        for col_info in columns_info:
            col_name = col_info['name']
            is_nullable = col_info['nullable'] == 'YES'
            
            try:
                # Check for NULL values
                self.cursor.execute(f"SELECT COUNT(*) FROM {table_name} WHERE {col_name} IS NULL")
                null_count = self.cursor.fetchone()[0]
                
                # Check for empty strings (for string columns)
                empty_count = 0
                if 'char' in col_info['type'].lower() or 'text' in col_info['type'].lower():
                    self.cursor.execute(f"SELECT COUNT(*) FROM {table_name} WHERE {col_name} = ''")
                    empty_count = self.cursor.fetchone()[0]
                
                # Get distinct count
                self.cursor.execute(f"SELECT COUNT(DISTINCT {col_name}) FROM {table_name}")
                distinct_count = self.cursor.fetchone()[0]
                
                # Calculate percentages
                null_pct = (null_count / row_count) * 100 if row_count > 0 else 0
                empty_pct = (empty_count / row_count) * 100 if row_count > 0 else 0
                cardinality = (distinct_count / row_count) * 100 if row_count > 0 else 0
                
                # Report findings
                status = "✓"
                notes = []
                
                if not is_nullable and null_count > 0:
                    status = "⚠️"
                    notes.append(f"{null_count} NULLs in non-nullable column")
                    quality_issues.append(f"Column '{col_name}': NULL constraint violation")
                
                if null_pct > 50:
                    status = "⚠️"
                    notes.append(f"High NULL rate: {null_pct:.1f}%")
                
                if empty_pct > 20:
                    status = "⚠️"
                    notes.append(f"High empty string rate: {empty_pct:.1f}%")
                
                if cardinality < 1 and row_count > 1:
                    status = "⚠️"
                    notes.append("All values identical")
                
                note_text = "; ".join(notes) if notes else f"Distinct: {distinct_count} ({cardinality:.1f}%)"
                print(f"{status} {col_name:<25} {note_text}")
                
            except Exception as e:
                print(f"✗ {col_name:<25} Error analyzing: {e}")
        
        if quality_issues:
            print(f"\n⚠️  Data Quality Issues Found:")
            for issue in quality_issues:
                print(f"   - {issue}")
    
    

    def check_unused_tables(self, specific_date_str: Optional[str] = None):
        """Check for potentially unused tables with enhanced reporting."""
        print(f"\n{'='*20} UNUSED TABLES ANALYSIS {'='*20}")
        
        specific_date = None
        if specific_date_str:
            try:
                specific_date = datetime.datetime.strptime(specific_date_str, '%Y-%m-%d').date()
                print(f"📅 Checking tables not used since: {specific_date}")
            except ValueError:
                print("⚠️  Invalid date format. Using all-time check.")
                specific_date_str = None
        else:
            print("📅 Performing all-time unused table check")

        unused_tables = []
        warnings = []
        
        try:
            if self.db_type == "sqlserver":
                query = """
                    SELECT
                        t.name AS TableName,
                        MAX(s.last_user_seek) AS LastSeek,
                        MAX(s.last_user_scan) AS LastScan,
                        MAX(s.last_user_lookup) AS LastLookup,
                        MAX(s.last_user_update) AS LastUpdate
                    FROM
                        sys.tables t
                    LEFT JOIN
                        sys.indexes i ON t.object_id = i.object_id
                    LEFT JOIN
                        sys.dm_db_index_usage_stats s ON i.object_id = s.object_id AND i.index_id = s.index_id
                    WHERE
                        t.is_ms_shipped = 0
                    GROUP BY
                        t.name
                    HAVING
                        (MAX(s.last_user_seek) IS NULL AND MAX(s.last_user_scan) IS NULL 
                         AND MAX(s.last_user_lookup) IS NULL AND MAX(s.last_user_update) IS NULL)
                """
                if specific_date:
                    query += f" OR (COALESCE(MAX(s.last_user_seek), MAX(s.last_user_scan), MAX(s.last_user_lookup), MAX(s.last_user_update)) < '{specific_date_str}')"

                self.cursor.execute(query)
                results = self.cursor.fetchall()
                for row in results:
                    unused_tables.append({
                        'name': row[0],
                        'last_activity': 'Never' if not any(row[1:]) else max([d for d in row[1:] if d])
                    })

            elif self.db_type == "postgresql":
                self.cursor.execute("""
                    SELECT
                        relname AS table_name,
                        last_seq_scan,
                        last_idx_scan,
                        n_tup_ins + n_tup_upd + n_tup_del AS total_modifications
                    FROM
                        pg_stat_user_tables
                    WHERE
                        schemaname = 'public'
                """)
                
                for row in self.cursor.fetchall():
                    table_name = row[0]
                    last_seq_scan = row[1]
                    last_idx_scan = row[2]
                    total_mods = row[3] or 0

                    is_unused = False
                    last_activity = None
                    
                    if not last_seq_scan and not last_idx_scan:
                        if total_mods == 0:
                            is_unused = True
                            last_activity = "Never"
                        else:
                            last_activity = "Modified but never read"
                    else:
                        last_activity = max([d for d in [last_seq_scan, last_idx_scan] if d])
                        if specific_date and last_activity and last_activity.date() < specific_date:
                            is_unused = True

                    if is_unused:
                        unused_tables.append({
                            'name': table_name,
                            'last_activity': last_activity or 'Never'
                        })

            elif self.db_type == "mysql":
                warnings.append("MySQL doesn't provide reliable table access statistics")
                warnings.append("Showing tables based on UPDATE_TIME (data modifications only)")
                
                query = """
                    SELECT TABLE_NAME, UPDATE_TIME, CREATE_TIME 
                    FROM INFORMATION_SCHEMA.TABLES 
                    WHERE TABLE_SCHEMA = DATABASE()
                """
                if specific_date:
                    query += f" AND (UPDATE_TIME IS NULL OR UPDATE_TIME < '{specific_date_str}')"
                    
                self.cursor.execute(query)
                for row in self.cursor.fetchall():
                    table_name = row[0]
                    update_time = row[1]
                    create_time = row[2]
                    
                    if not update_time:
                        last_activity = f"Created: {create_time}, Never updated"
                    else:
                        last_activity = f"Last updated: {update_time}"
                    
                    unused_tables.append({
                        'name': table_name,
                        'last_activity': last_activity
                    })

            # Report results
            if warnings:
                print("⚠️  Warnings:")
                for warning in warnings:
                    print(f"   - {warning}")
                print()

            if unused_tables:
                print(f"📋 Found {len(unused_tables)} potentially unused tables:")
                print(f"{'Table Name':<30} {'Last Activity':<40}")
                print("-" * 75)
                for table in unused_tables:
                    activity = str(table['last_activity'])[:39]
                    print(f"{table['name']:<30} {activity:<40}")
            else:
                print("✅ No unused tables found based on the criteria.")

        except Exception as e:
            print(f"✗ Error checking unused tables: {e}")


    
    def export_schema_report(self):
        """Export a comprehensive schema report to file."""
        try:
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"schema_report_{self.db_type}_{timestamp}.xlsx"

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Database Report"

            # Header section
            ws["A1"] = "DATABASE SCHEMA REPORT"
            ws["A1"].font = Font(size=14, bold=True)
            ws["A2"] = f"Generated: {datetime.datetime.now()}"
            ws["A3"] = f"Database Type: {self.db_type.upper()}"

            # Basic stats
            tables = self.get_tables()
            ws["A5"] = "Total Tables:"
            ws["B5"] = len(tables)

            ws["A6"] = "Total Views:"
            ws["B6"] = self._get_views_count()

            ws["A7"] = "Total Schemas:"
            ws["B7"] = self._get_schemas_count()

            # Section Title
            ws["A9"] = "TABLES OVERVIEW"
            ws["A9"].font = Font(size=12, bold=True)

            # Table headers
            ws["A10"] = "Schema"
            ws["B10"] = "Table Name"
            ws["C10"] = "Row Count"

            for col in ["A", "B", "C"]:
                ws[f"{col}10"].font = Font(bold=True)
                ws[f"{col}10"].alignment = Alignment(horizontal="center")

            # Fill table data
            row_num = 11
            for schema, table in tables:
                try:
                    self.cursor.execute(f"SELECT COUNT(*) FROM {schema}.{table}")
                    row_count = self.cursor.fetchone()[0]
                    ws[f"A{row_num}"] = schema
                    ws[f"B{row_num}"] = table
                    ws[f"C{row_num}"] = row_count
                except:
                    ws[f"A{row_num}"] = schema
                    ws[f"B{row_num}"] = table
                    ws[f"C{row_num}"] = "Error"
                row_num += 1

            # Auto-adjust column width
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2

            # Save to Excel
            wb.save(filename)
                
            print(f"✅ Schema report exported to: {filename}")
            
        except Exception as e:
            print(f"✗ Error exporting report: {e}")
    
    def detect_duplicate_rows(self, schema: str, table_name: str):
        """Detect and analyze duplicate rows with fuzzy matching for similar content."""
        print(f"\n{'='*20} DUPLICATE DETECTION: {schema}.{table_name} {'='*20}")
        
        try:
            # Get table structure
            columns_info = self._get_column_info(schema, table_name)
            if not columns_info:
                print("❌ Could not retrieve table structure.")
                return
            
            # Get row count
            self.cursor.execute(f"SELECT COUNT(*) FROM {schema}.{table_name}")
            total_rows = self.cursor.fetchone()[0]
            
            if total_rows == 0:
                print("ℹ️  Table is empty.")
                return
            
            if total_rows > 10000:
                proceed = input(f"⚠️  Large table ({total_rows:,} rows). This may take time. Continue? (y/N): ").strip().lower()
                if proceed != 'y':
                    return
            
            print(f"📊 Analyzing {total_rows:,} rows across {len(columns_info)} columns...")
            
            # Show available columns for analysis
            print(f"\n📋 Available columns:")
            text_columns = []
            numeric_columns = []
            
            for i, col in enumerate(columns_info, 1):
                col_type = col['type'].lower()
                if any(t in col_type for t in ['char', 'text', 'varchar', 'string']):
                    text_columns.append(col['name'])
                    print(f"{i:3d}. {col['name']:<25} ({col['type']}) - TEXT")
                elif any(t in col_type for t in ['int', 'decimal', 'numeric', 'float', 'double', 'real']):
                    numeric_columns.append(col['name'])
                    print(f"{i:3d}. {col['name']:<25} ({col['type']}) - NUMERIC")
                else:
                    print(f"{i:3d}. {col['name']:<25} ({col['type']}) - OTHER")
            
            print(f"\n🔧 Duplicate Detection Options:")
            print("1. Exact duplicate rows (all columns)")
            print("2. Fuzzy matching on specific text columns")
            print("3. Combination matching (exact + fuzzy)")
            print("4. Custom column selection")
            
            detection_type = input("\nChoose detection method (1-4): ").strip()
            
            duplicates_found = []
            
            if detection_type == "1":
                duplicates_found = self._find_exact_duplicates(schema, table_name, columns_info)
                
            elif detection_type == "2":
                if not text_columns:
                    print("❌ No text columns found for fuzzy matching.")
                    return
                duplicates_found = self._find_fuzzy_duplicates(schema, table_name, text_columns)
                
            elif detection_type == "3":
                duplicates_found = self._find_combination_duplicates(schema, table_name, columns_info, text_columns)
                
            elif detection_type == "4":
                selected_columns = self._get_user_column_selection(columns_info)
                if not selected_columns:
                    return
                duplicates_found = self._find_custom_duplicates(schema, table_name, selected_columns, columns_info)
            
            else:
                print("❌ Invalid selection.")
                return
            
            # Display results
            if duplicates_found:
                self._display_duplicate_results(duplicates_found, schema, table_name)
                
                # Offer to export to Excel
                export_choice = input(f"\n💾 Export {len(duplicates_found)} duplicate groups to Excel? (y/N): ").strip().lower()
                if export_choice == 'y':
                    self._export_duplicates_to_excel(duplicates_found, schema, table_name, detection_type)
            else:
                print("✅ No duplicates found!")
                
        except Exception as e:
            print(f"❌ Error detecting duplicates: {e}")
    
    def _find_exact_duplicates(self, schema: str, table_name: str, columns_info: List[Dict]) -> List[Dict]:
        """Find exact duplicate rows."""
        print("🔍 Finding exact duplicates...")
        
        # Create column list for query
        columns = [col['name'] for col in columns_info]
        column_list = ', '.join(columns)
        
        # Find duplicates using GROUP BY and HAVING
        query = f"""
        SELECT {column_list}, COUNT(*) as duplicate_count
        FROM {schema}.{table_name}
        GROUP BY {column_list}
        HAVING COUNT(*) > 1
        ORDER BY duplicate_count DESC
        """
        
        self.cursor.execute(query)
        results = self.cursor.fetchall()
        
        duplicates = []
        for row in results:
            duplicate_count = row[-1]
            row_data = dict(zip(columns, row[:-1]))
            duplicates.append({
                'type': 'exact',
                'count': duplicate_count,
                'data': row_data,
                'similarity': 100.0,
                'similar_rows': [row_data] * duplicate_count
            })
        
        return duplicates
    
    def _find_fuzzy_duplicates(self, schema: str, table_name: str, text_columns: List[str], similarity_threshold: float = 0.7) -> List[Dict]:
        """Find fuzzy duplicates based on text similarity."""
        print(f"🔍 Finding fuzzy duplicates (similarity >= {similarity_threshold*100:.0f}%)...")
        
        # Get similarity threshold from user
        threshold_input = input(f"Enter similarity threshold (0.1-0.9, default {similarity_threshold}): ").strip()
        if threshold_input:
            try:
                similarity_threshold = float(threshold_input)
                similarity_threshold = max(0.1, min(0.9, similarity_threshold))
            except ValueError:
                pass
        
        # Select which text column to analyze
        if len(text_columns) > 1:
            print(f"\nText columns available:")
            for i, col in enumerate(text_columns, 1):
                print(f"{i}. {col}")
            
            try:
                col_choice = int(input(f"Select column number (1-{len(text_columns)}): ")) - 1
                if 0 <= col_choice < len(text_columns):
                    target_column = text_columns[col_choice]
                else:
                    target_column = text_columns[0]
            except ValueError:
                target_column = text_columns[0]
        else:
            target_column = text_columns[0]
        
        print(f"📝 Analyzing column: {target_column}")
        
        # Get all unique values from the target column
        self.cursor.execute(f"SELECT DISTINCT {target_column} FROM {schema}.{table_name} WHERE {target_column} IS NOT NULL ORDER BY {target_column}")
        unique_values = [row[0] for row in self.cursor.fetchall()]
        
        if len(unique_values) > 1000:
            print(f"⚠️  Large dataset ({len(unique_values)} unique values). This may take time...")
        
        # Find similar groups
        processed = set()
        duplicate_groups = []
        
        for i, value1 in enumerate(unique_values):
            if value1 in processed:
                continue
                
            similar_group = [value1]
            processed.add(value1)
            
            for j, value2 in enumerate(unique_values[i+1:], i+1):
                if value2 in processed:
                    continue
                    
                similarity = self._calculate_similarity(str(value1), str(value2))
                if similarity >= similarity_threshold:
                    similar_group.append(value2)
                    processed.add(value2)
            
            if len(similar_group) > 1:
                # Get count for each similar value
                group_data = []
                total_count = 0
                for val in similar_group:
                    if self.db_type == 'sqlserver':
                        self.cursor.execute(f"SELECT COUNT(*) FROM {schema}.{table_name} WHERE {target_column} = ?", (val,))
                    else:
                        self.cursor.execute(f"SELECT COUNT(*) FROM {schema}.{table_name} WHERE {target_column} = %s", (val,))
                    count = self.cursor.fetchone()[0]
                    group_data.append({'value': val, 'count': count})
                    total_count += count
                
                duplicate_groups.append({
                    'type': 'fuzzy',
                    'column': target_column,
                    'count': total_count,
                    'similarity': similarity_threshold * 100,
                    'similar_values': group_data,
                    'group_size': len(similar_group)
                })
        
        return duplicate_groups
    
    def _find_combination_duplicates(self, schema: str, table_name: str, columns_info: List[Dict], text_columns: List[str]) -> List[Dict]:
        """Find duplicates using combination of exact and fuzzy matching."""
        print("🔍 Finding combination duplicates (exact + fuzzy)...")
        
        # First find exact duplicates
        exact_duplicates = self._find_exact_duplicates(schema, table_name, columns_info)
        
        # Then find fuzzy duplicates
        fuzzy_duplicates = []
        if text_columns:
            fuzzy_duplicates = self._find_fuzzy_duplicates(schema, table_name, text_columns, 0.8)
        
        # Combine results
        all_duplicates = []
        all_duplicates.extend(exact_duplicates)
        all_duplicates.extend(fuzzy_duplicates)
        
        return all_duplicates
    
    def _find_custom_duplicates(self, schema: str, table_name: str, selected_columns: List[str], columns_info: List[Dict]) -> List[Dict]:
        """Find duplicates based on user-selected columns."""
        print(f"🔍 Finding duplicates based on selected columns: {', '.join(selected_columns)}")
        
        column_list = ', '.join(selected_columns)
        
        query = f"""
        SELECT {column_list}, COUNT(*) as duplicate_count
        FROM {schema}.{table_name}
        GROUP BY {column_list}
        HAVING COUNT(*) > 1
        ORDER BY duplicate_count DESC
        """
        
        self.cursor.execute(query)
        results = self.cursor.fetchall()
        
        duplicates = []
        for row in results:
            duplicate_count = row[-1]
            row_data = dict(zip(selected_columns, row[:-1]))
            duplicates.append({
                'type': 'custom',
                'count': duplicate_count,
                'data': row_data,
                'columns': selected_columns,
                'similar_rows': [row_data] * duplicate_count
            })
        
        return duplicates
    
    def _get_user_column_selection(self, columns_info: List[Dict]) -> List[str]:
        """Get user selection of columns for duplicate detection."""
        print(f"\nSelect columns for duplicate detection (comma-separated numbers):")
        for i, col in enumerate(columns_info, 1):
            print(f"{i:3d}. {col['name']} ({col['type']})")
        
        try:
            selection = input("Enter column numbers (e.g., 1,3,5): ").strip()
            indices = [int(x.strip()) - 1 for x in selection.split(',')]
            selected_columns = [columns_info[i]['name'] for i in indices if 0 <= i < len(columns_info)]
            
            if not selected_columns:
                print("❌ No valid columns selected.")
                return []
                
            print(f"✅ Selected columns: {', '.join(selected_columns)}")
            return selected_columns
            
        except (ValueError, IndexError):
            print("❌ Invalid selection.")
            return []
    
    def _calculate_similarity(self, str1: str, str2: str) -> float:
        """Calculate similarity between two strings using multiple methods."""
        if not str1 or not str2:
            return 0.0
        
        # Clean strings for comparison
        clean_str1 = re.sub(r'[^\w\s]', '', str1.lower().strip())
        clean_str2 = re.sub(r'[^\w\s]', '', str2.lower().strip())
        
        # Exact match
        if clean_str1 == clean_str2:
            return 1.0
        
        # Sequence matcher (overall similarity)
        seq_similarity = SequenceMatcher(None, clean_str1, clean_str2).ratio()
        
        # Check if one string contains the other (substring)
        if clean_str1 in clean_str2 or clean_str2 in clean_str1:
            substring_bonus = 0.2
        else:
            substring_bonus = 0.0
        
        # Word-level similarity
        words1 = set(clean_str1.split())
        words2 = set(clean_str2.split())
        if words1 and words2:
            word_similarity = len(words1.intersection(words2)) / len(words1.union(words2))
        else:
            word_similarity = 0.0
        
        # Combined similarity score
        final_similarity = max(seq_similarity, word_similarity) + substring_bonus
        return min(1.0, final_similarity)
    
    def _display_duplicate_results(self, duplicates: List[Dict], schema: str, table_name: str):
        """Display duplicate detection results."""
        print(f"\n🎯 DUPLICATE DETECTION RESULTS FOR {schema}.{table_name}:")
        print("=" * 70)
        
        total_duplicate_rows = 0
        
        for i, dup in enumerate(duplicates, 1):
            print(f"\n📝 Group {i}: {dup['type'].upper()} duplicates")
            print("-" * 50)
            
            if dup['type'] == 'exact':
                print(f"Count: {dup['count']} identical rows")
                print("Data:")
                for key, value in dup['data'].items():
                    print(f"  {key}: {value}")
                total_duplicate_rows += dup['count']
                
            elif dup['type'] == 'fuzzy':
                print(f"Similar values in column '{dup['column']}':")
                print(f"Total rows affected: {dup['count']}")
                print(f"Similarity threshold: {dup.get('similarity', 70):.1f}%")
                for val_info in dup['similar_values']:
                    print(f"  '{val_info['value']}' ({val_info['count']} rows)")
                total_duplicate_rows += dup['count']
                
            elif dup['type'] == 'custom':
                print(f"Count: {dup['count']} rows with matching {', '.join(dup['columns'])}")
                print("Matching data:")
                for key, value in dup['data'].items():
                    print(f"  {key}: {value}")
                total_duplicate_rows += dup['count']
        
        print(f"\n📊 SUMMARY:")
        print(f"Total duplicate groups found: {len(duplicates)}")
        print(f"Total rows involved in duplicates: {total_duplicate_rows:,}")
    
    def _export_duplicates_to_excel(self, duplicates: List[Dict], schema: str, table_name: str, detection_type: str):
        """Export duplicate detection results to Excel file."""
        try:
            
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Duplicates Summary"
            
            # Add header
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws['A1'] = f"Duplicate Analysis Report - {schema}.{table_name}"
            ws['A1'].font = Font(size=14, bold=True)
            ws['A2'] = f"Generated: {timestamp}"
            ws['A3'] = f"Detection Method: {detection_type}"
            
            # Summary section
            row_num = 5
            ws[f'A{row_num}'] = "SUMMARY"
            ws[f'A{row_num}'].font = Font(size=12, bold=True)
            ws[f'A{row_num}'].fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
            
            row_num += 1
            ws[f'A{row_num}'] = "Group"
            ws[f'B{row_num}'] = "Type"
            ws[f'C{row_num}'] = "Row Count"
            ws[f'D{row_num}'] = "Description"
            
            # Style header row
            for col in ['A', 'B', 'C', 'D']:
                cell = ws[f'{col}{row_num}']
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            
            # Add duplicate groups
            for i, dup in enumerate(duplicates, 1):
                row_num += 1
                ws[f'A{row_num}'] = i
                ws[f'B{row_num}'] = dup['type'].upper()
                ws[f'C{row_num}'] = dup['count']
                
                if dup['type'] == 'exact':
                    description = f"Exact duplicates: {', '.join([f'{k}={v}' for k, v in list(dup['data'].items())[:2]])}"
                elif dup['type'] == 'fuzzy':
                    description = f"Similar values in '{dup['column']}': {len(dup['similar_values'])} variants"
                else:
                    description = f"Custom duplicates on: {', '.join(dup.get('columns', []))}"
                
                ws[f'D{row_num}'] = description[:100]  # Truncate long descriptions
            
            # Create detailed sheets for each duplicate group
            for i, dup in enumerate(duplicates, 1):
                if i > 10:  # Limit to prevent too many sheets
                    break
                    
                sheet_name = f"Group_{i}_{dup['type'][:10]}"
                detail_ws = wb.create_sheet(title=sheet_name)
                
                # Add group details
                detail_ws['A1'] = f"Duplicate Group {i} - {dup['type'].upper()}"
                detail_ws['A1'].font = Font(size=12, bold=True)
                
                if dup['type'] == 'fuzzy' and 'similar_values' in dup:
                    # Fuzzy duplicates details
                    detail_ws['A3'] = "Value"
                    detail_ws['B3'] = "Count"
                    detail_ws['A3'].font = Font(bold=True)
                    detail_ws['B3'].font = Font(bold=True)
                    
                    row = 4
                    for val_info in dup['similar_values']:
                        detail_ws[f'A{row}'] = str(val_info['value'])
                        detail_ws[f'B{row}'] = val_info['count']
                        row += 1
                
                elif dup['type'] in ['exact', 'custom'] and 'data' in dup:
                    # Exact/custom duplicates details
                    detail_ws['A3'] = "Column"
                    detail_ws['B3'] = "Value"
                    detail_ws['A3'].font = Font(bold=True)
                    detail_ws['B3'].font = Font(bold=True)
                    
                    row = 4
                    for key, value in dup['data'].items():
                        detail_ws[f'A{row}'] = key
                        detail_ws[f'B{row}'] = str(value)
                        row += 1
            
            # Auto-adjust column widths
            for ws_obj in wb.worksheets:
                for column in ws_obj.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_obj.column_dimensions[column_letter].width = adjusted_width
            
            # Save file
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"duplicates_{schema}.{table_name}_{detection_type}_{timestamp}.xlsx"
            wb.save(filename)
            
            print(f"✅ Duplicate report exported to: {filename}")
            print(f"📊 Report contains {len(duplicates)} duplicate groups across {len(wb.worksheets)} sheets")
            
        except Exception as e:
            print(f"❌ Error exporting to Excel: {e}")
            print("💡 Make sure you have 'openpyxl' installed: pip install openpyxl")



    def _get_view_definition(self, schema: str, view_name: str) -> Optional[str]:
        """Return the SQL definition of a view, or None if not found."""
        try:
            if self.db_type.lower() == "sqlserver":
                # sys.sql_modules holds object definitions
                sql = """
                    SELECT m.definition
                    FROM sys.views v
                    JOIN sys.schemas s ON v.schema_id = s.schema_id
                    JOIN sys.sql_modules m ON v.object_id = m.object_id
                    WHERE s.name = ? AND v.name = ?
                """
                self.cursor.execute(sql, (schema, view_name))
                row = self.cursor.fetchone()
                return row[0] if row else None

            elif self.db_type.lower() == "postgresql":
                # use pg_get_viewdef for better formatting
                sql = "SELECT pg_get_viewdef(quote_ident(%s) || '.' || quote_ident(%s)::regclass, true)"
                # simpler: fetch from information_schema.views
                self.cursor.execute("""
                    SELECT view_definition
                    FROM information_schema.views
                    WHERE table_schema = %s AND table_name = %s
                """, (schema, view_name))
                row = self.cursor.fetchone()
                return row[0] if row else None

            elif self.db_type.lower() == "mysql":
                self.cursor.execute("""
                    SELECT VIEW_DEFINITION
                    FROM INFORMATION_SCHEMA.VIEWS
                    WHERE TABLE_SCHEMA = %s AND TABLE_NAME = %s
                """, (schema, view_name))
                row = self.cursor.fetchone()
                return row[0] if row else None

            else:
                return None
        except Exception:
            return None


    def _is_view(self, schema: str, name: str) -> bool:
        """Return True if given schema.name is a view in the DB."""
        try:
            if self.db_type.lower() == "sqlserver":
                self.cursor.execute("""
                    SELECT 1 FROM INFORMATION_SCHEMA.VIEWS WHERE TABLE_SCHEMA = ? AND TABLE_NAME = ?
                """, (schema, name))
                return bool(self.cursor.fetchone())

            else:  # postgres & mysql use %s
                self.cursor.execute("""
                    SELECT 1 FROM INFORMATION_SCHEMA.VIEWS WHERE TABLE_SCHEMA = %s AND TABLE_NAME = %s
                """, (schema, name))
                return bool(self.cursor.fetchone())
        except Exception:
            return False


    def _parse_sql_references(self, sql_text: str) -> Set[Tuple[Optional[str], str]]:
        """
        Heuristically parse SQL and return a set of (schema_or_None, object_name).
        Looks for tokens after FROM, JOIN, INTO, UPDATE, etc.
        Returns names as they appear: either 'schema.table' will be split,
        or plain 'table' will be returned with schema=None.
        """
        if not sql_text:
            return set()

        # Remove string literals and comments to avoid false positives
        sql_no_strings = re.sub(r"(--[^\n]*\n)|('(?:''|[^'])*')|(\"(?:\"\"|[^\"])*\")|(/\*.*?\*/)", " ", sql_text, flags=re.S)

        # Find candidate object names after FROM, JOIN, UPDATE, INTO, APPLY
        pattern = re.compile(r'\b(?:FROM|JOIN|INTO|UPDATE|MERGE|APPLY)\s+([`"\[]?)([A-Za-z0-9_\.\$-]+)\1', flags=re.IGNORECASE)
        matches = pattern.findall(sql_no_strings)

        refs = set()
        for _, fullname in matches:
            # strip trailing commas or parentheses
            fullname = fullname.strip().rstrip(',').rstrip(')')
            # if alias like schema.table AS t, regex above stops at whitespace so okay
            if '.' in fullname:
                parts = fullname.split('.')
                # support quoted-like [schema].[table] or schema.table
                if len(parts) >= 2:
                    schema = parts[-2].strip('[]"`')
                    name = parts[-1].strip('[]"`')
                    refs.add((schema, name))
            else:
                # unqualified name: return as (None, name)
                refs.add((None, fullname.strip('[]"`')))

        # Exclude common CTE names declared by WITH ... AS (they are not real objects)
        cte_pattern = re.compile(r'WITH\s+([A-Za-z0-9_]+)\s+AS', flags=re.IGNORECASE)
        ctes = set(m.group(1).lower() for m in cte_pattern.finditer(sql_text))
        refs = {(s, n) for (s, n) in refs if (n.lower() not in ctes)}

        return refs


    def build_view_hierarchy(self,
                            root_schema: str,
                            root_view: str,
                            output_path: str = "view_hierarchy",
                            output_format: str = "png",
                            max_depth: int = 10) -> str:
        """
        Build dependency graph for a view and save to output_path (without extension).
        Returns path to generated file (e.g., 'view_hierarchy.png').

        Note: uses graphviz if available; fallback to networkx+matplotlib.
        """
        try:
            from graphviz import Digraph
            use_graphviz = True
        except Exception:
            use_graphviz = False

        # Helper to normalize names for node ids
        def node_id(schema, name):
            return f"{schema}.{name}" if schema else name

        graph_edges = []  # list of (parent, child)
        nodes = set()
        visited = set()

        # queue for DFS/BFS: tuples (schema, name, depth)
        stack = [(root_schema, root_view, 0)]

        while stack:
            schema, name, depth = stack.pop()
            key = (schema or '', name.lower())
            if key in visited or depth > max_depth:
                continue
            visited.add(key)

            this_node = node_id(schema, name)
            nodes.add(this_node)

            # Get view definition (if view) otherwise try to skip fetching for tables
            definition = None
            try:
                definition = self._get_view_definition(schema, name)
            except Exception:
                definition = None

            # If it's a view and we could get definition, parse for dependencies
            if definition:
                refs = self._parse_sql_references(definition)
                for ref_schema, ref_name in refs:
                    # resolve schema: if None, assume same schema as view (best-effort)
                    resolved_schema = ref_schema or schema
                    child = node_id(resolved_schema, ref_name)
                    graph_edges.append((this_node, child))
                    nodes.add(child)

                    # If referenced object is itself a view, recurse
                    try:
                        if self._is_view(resolved_schema, ref_name):
                            stack.append((resolved_schema, ref_name, depth + 1))
                    except Exception:
                        # If check fails, still add node but don't recurse
                        pass
            else:
                # If we couldn't fetch definition, optionally try to detect dependencies using information_schema (rare)
                pass

        # Render graph
        output_file = f"{output_path}.{output_format}"
        if use_graphviz:
            dot = Digraph(comment=f"Dependency graph for {node_id(root_schema, root_view)}")
            # add nodes
            for n in nodes:
                dot.node(n, n)
            for a, b in graph_edges:
                dot.edge(a, b)
            try:
                dot.render(filename=output_path, format=output_format, cleanup=True)
                return output_file
            except Exception as e:
                # fallback if graphviz binary missing or render failed
                use_graphviz = False

        # Fallback: networkx + matplotlib
        try:
            import networkx as nx
            import matplotlib.pyplot as plt
        except Exception as e:
            raise RuntimeError("Neither graphviz nor networkx+matplotlib are available to render the graph.") from e

        G = nx.DiGraph()
        for n in nodes:
            G.add_node(n)
        for a, b in graph_edges:
            G.add_edge(a, b)

        plt.figure(figsize=(12, max(6, len(nodes) * 0.3)))
        try:
            pos = nx.nx_agraph.graphviz_layout(G, prog="dot")
        except Exception:
            pos = nx.spring_layout(G)
        nx.draw(G, pos, with_labels=True, arrows=True, node_size=2000, font_size=8)
        plt.tight_layout()
        plt.savefig(output_file, dpi=150)
        plt.close()
        return output_file







def main():
    """Main application loop with enhanced menu system."""
    print("🗄️  Enhanced Database Schema Analyzer")
    print("=" * 50)
    print("Supported databases: PostgreSQL, SQL Server, MySQL")
    
    db_type = input("\nEnter database type (postgresql/sqlserver/mysql): ").strip().lower()
    
    if db_type not in ["postgresql", "sqlserver", "mysql"]:
        print("❌ Unsupported database type.")
        return
    
    analyzer = DatabaseAnalyzer()
    
    # Get connection parameters and connect
    params = analyzer.get_connection_params(db_type)
    if not analyzer.connect_database(db_type, params):
        return
    
    try:
        # Show database overview
        analyzer.get_database_overview()
        
        while True:
            print("\n" + "="*50)
            print("🔧 AVAILABLE ACTIONS:")
            print("="*50)
            print("1. 📈 Database overview")
            print("2. 📋 List all Tables")
            print("3. 👁  List all VIEWS") 
            print("4. 🔍 List indexes for a specific table")
            print("5. 📊 Get detailed table analysis")
            print("6. 🚫 Check for unused tables")
            print("7. 🔄 Detect duplicate rows")
            print("8. 🔍 Detect similar tables")
            print("9. 🔎 Find tables by column name")
            print("10. 📄 Export schema report")
            print("11. ❌ Exit")

            choice = input("\nEnter your choice (1-10): ").strip()

            if choice == "1":
                analyzer.get_database_overview()

            elif choice == "2":
                print("\n📋 LISTING ALL TABLES:")
                print("-" * 40)
                tables = analyzer.get_tables()
                if tables:
                    # Get row counts and column info for each table
                    table_stats = []
                    for schema, table_name in tables:
                        # Use schema-qualified name for querying row count
                        qualified_name = f"{schema}.{table_name}"
                        analyzer.cursor.execute(f"SELECT COUNT(*) FROM {qualified_name}")
                        row_count = analyzer.cursor.fetchone()[0]

                        # Get column info using schema and table name
                        columns_info = analyzer._get_column_info(schema, table_name)

                        # Store stats
                        table_stats.append((qualified_name, row_count, len(columns_info)))

                    # Sort by row count descending
                    table_stats.sort(key=lambda x: x[1], reverse=True)

                    # Get max length of table names for spacing
                    max_len = max(len(name) for name, _, _ in table_stats)

                    # Print header
                    print(f"{'No.':<4} {'Tables':<{max_len}}   {'Rows':>6}   {'Columns':>7}")
                    print('-' * (max_len + 25))

                    # Print each table's stats
                    for i, (qualified_name, row_count, col_count) in enumerate(table_stats, 1):
                        print(f"{i:<4} {qualified_name:<{max_len}}   {row_count:>6}   {col_count:>7}")

                else:
                    print("ℹ️  No tables found in the database.")


            elif choice == "3":
                print("\n📋 LISTING ALL VIEWS:")
                print("-" * 40)
                views = analyzer.get_views()
                if views:
                    view_stats = []
                    for schema, view_name in views:
                        qualified_name = f"{schema}.{view_name}"

                        # Try to get row count (may be expensive for views). Fail gracefully.
                        try:
                            analyzer.cursor.execute(f"SELECT COUNT(*) FROM {qualified_name}")
                            row_count = analyzer.cursor.fetchone()[0]
                        except Exception:
                            row_count = None  # Could not compute (complex view, permission, etc.)

                        columns_info = analyzer._get_column_info(schema, view_name)
                        col_count = len(columns_info)

                        view_stats.append((qualified_name, row_count, col_count, schema, view_name))

                    # Put None counts at the end when sorting descending (treat None as -1)
                    view_stats.sort(key=lambda x: (x[1] is None, -(x[1] or 0)))

                    max_len = max(len(item[0]) for item in view_stats)

                    # Print header
                    print(f"{'No.':<4} {'Views':<{max_len}}   {'Rows':>8}   {'Columns':>8}")
                    print('-' * (max_len + 30))

                    for i, (qualified_name, row_count, col_count, schema, view_name) in enumerate(view_stats, 1):
                        rows_display = f"{row_count:,}" if row_count is not None else "N/A"
                        print(f"{i:<4} {qualified_name:<{max_len}}   {rows_display:>8}   {col_count:>8}")

                    
                    while True:
                        try:
                            print("\n🔍 Select VIEW to see Dependency Hierarchy:")
                            print("-" * 45)
                            view_choice = input(f"\nEnter View number (1-{len(view_stats)}) or 0 to go back: ")
                            view_choice = int(view_choice)
                            if view_choice == 0:
                                break
                            if 1 <= view_choice <= len(view_stats):
                                # Use the sorted view_stats list instead of original views list
                                selected_schema = view_stats[view_choice - 1][3]  # schema
                                selected_view = view_stats[view_choice - 1][4]    # view_name
                                output = analyzer.build_view_hierarchy(selected_schema, selected_view, output_path=f"{selected_schema}_{selected_view}", output_format="png")
                                print(f"{selected_schema}_{selected_view} View hierarchy image saved to:", output)
                                break

                            else:
                                print(f"❌ Invalid number. Please enter 1-{len(view_stats)} or 0.")
                        except ValueError:
                            print("❌ Invalid input. Please enter a number.")

                else:
                    print("ℹ️  No views found in the database.")


            elif choice == "4":
                tables = analyzer.get_tables()
                if not tables:
                    print("ℹ️  No tables available to analyze indexes.")
                    continue

                print("\n🔍 SELECT TABLE FOR INDEX ANALYSIS:")
                print("-" * 45)
                for i, (schema, table_name) in enumerate(tables, 1):
                    print(f"{i:3d}. {schema}.{table_name}")

                while True:
                    try:
                        table_choice = input(f"\nEnter table number (1-{len(tables)}) or 0 to go back: ")
                        table_choice = int(table_choice)
                        if table_choice == 0:
                            break
                        if 1 <= table_choice <= len(tables):
                            schema, selected_table = tables[table_choice - 1]
                            analyzer.list_indexes_for_table(schema, selected_table)
                            break
                        else:
                            print(f"❌ Invalid number. Please enter 1-{len(tables)} or 0.")
                    except ValueError:
                        print("❌ Invalid input. Please enter a number.")


            elif choice == "5":
                tables = analyzer.get_tables()
                if not tables:
                    print("No tables available to analyze.")
                    continue

                print("\nTABLE ANALYSIS OPTIONS:")
                print("-" * 30)
                print("1. Analyze single table")
                print("2. Export all tables analysis to Excel")
                
                analysis_choice = input("Choose option (1/2): ").strip()
                
                if analysis_choice == "1":
                    # Your existing single table analysis code
                    print("\nSELECT TABLE FOR DETAILED ANALYSIS:")
                    print("-" * 45)
                    for i, (schema, table_name) in enumerate(tables, 1):
                        print(f"{i:3d}. {schema}.{table_name}")

                    while True:
                        try:
                            table_choice = input(f"\nEnter table number (1-{len(tables)}) or 0 to go back: ")
                            table_choice = int(table_choice)
                            if table_choice == 0:
                                break
                            if 1 <= table_choice <= len(tables):
                                schema, selected_table = tables[table_choice - 1]
                                analyzer.get_table_details_and_quality(schema, selected_table)
                                break
                            else:
                                print(f"Invalid number. Please enter 1-{len(tables)} or 0.")
                        except ValueError:
                            print("Invalid input. Please enter a number.")
                            
                elif analysis_choice == "2":
                    analyzer.export_all_tables_analysis()
                else:
                    print("Invalid choice.")

                while True:
                    try:
                        table_choice = input(f"\nEnter table number (1-{len(tables)}) or 0 to go back: ")
                        table_choice = int(table_choice)
                        if table_choice == 0:
                            break
                        if 1 <= table_choice <= len(tables):
                            schema, selected_table = tables[table_choice - 1]
                            analyzer.get_table_details_and_quality(schema, selected_table)
                            break
                        else:
                            print(f"❌ Invalid number. Please enter 1-{len(tables)} or 0.")
                    except ValueError:
                        print("❌ Invalid input. Please enter a number.")


            elif choice == "6":
                print("\n🚫 UNUSED TABLES ANALYSIS:")
                print("-" * 30)
                date_input = input("Enter date (YYYY-MM-DD) to check tables unused since then\n(or press Enter for all-time check): ").strip()
                analyzer.check_unused_tables(date_input if date_input else None)


            elif choice == "7":
                tables = analyzer.get_tables()
                if not tables:
                    print("ℹ️  No tables available to check for duplicates.")
                    continue

                print("\n🔄 SELECT TABLE FOR DUPLICATE DETECTION:")
                print("-" * 45)
                for i, (schema, table_name) in enumerate(tables, 1):
                    print(f"{i:3d}. {schema}.{table_name}")

                while True:
                    try:
                        table_choice = input(f"\nEnter table number (1-{len(tables)}) or 0 to go back: ")
                        table_choice = int(table_choice)
                        if table_choice == 0:
                            break
                        if 1 <= table_choice <= len(tables):
                            schema, selected_table = tables[table_choice - 1]
                            analyzer.detect_duplicate_rows(schema, selected_table)
                            break
                        else:
                            print(f"❌ Invalid number. Please enter 1-{len(tables)} or 0.")
                    except ValueError:
                        print("❌ Invalid input. Please enter a number.")


            elif choice == "8":
                print("\n🔍 DETECTING SIMILAR TABLES:")
                print("-" * 35)
                analyzer.detect_similar_tables()


            elif choice == "9":
                print("\n🔎 FIND TABLES BY COLUMN NAME:")
                print("-" * 35)
                column_name = input("Enter column name to search for: ").strip()
                if column_name:
                    analyzer.find_tables_by_column(column_name)
                else:
                    print("❌ Column name cannot be empty.")


            elif choice == "10":
                print("\n📄 EXPORTING SCHEMA REPORT:")
                print("-" * 30)
                analyzer.export_schema_report()


            elif choice == "11":
                print("\n👋 Exiting application...")
                break

            else:
                print("❌ Invalid choice. Please enter a number from 1-11.")

    except KeyboardInterrupt:
        print("\n\n⚠️  Application interrupted by user.")
    except Exception as e:
        print(f"\n❌ Unexpected error: {e}")
    finally:
        analyzer.close_connection()

if __name__ == "__main__":
    main()
